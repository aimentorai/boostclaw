#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
选品 Excel 生成脚本（product_chart_get skill）

用法：
    python3 generate_excel.py <OUTPUT_DIR> <DS>

参数：
    OUTPUT_DIR  选品目录，如 ~/Desktop/20260328选品
    DS          数据分区日期，如 20260328

目录内需包含（二选一）：
    优先模式：merged_products.json（AI 在上下文中一次完成提取+预处理，推荐）
    兼容模式：data-get-第x页数据.json + ai_results.json（旧流程，自动 fallback）

生成文件：
    YYYYMMDD选品.xlsx
    YYYYMMDD选品复筛.xlsx
    已排查产品.xlsx  （追加，不覆盖，位于选品目录同级）
"""

import re
import json
import os
import sys
import glob
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime

# ===== 参数处理 =====
def get_default_workspace():
    """从环境变量 $WORKSPACE 读取，未设置则默认 ~/Desktop"""
    workspace = os.getenv('WORKSPACE')
    if workspace:
        return os.path.expanduser(workspace)
    return os.path.expanduser('~/Desktop')


def parse_args():
    """
    支持三种调用方式：
    1. 环境变量方式（推荐）: WORKSPACE=~/Desktop python3 generate_excel.py --date <YYYYMMDD>
    2. 新方式: generate_excel.py --workspace <WORKSPACE> --date <YYYYMMDD>
    3. 旧方式（向后兼容）: generate_excel.py <OUTPUT_DIR> <DS> [CUSTOM_JSON]
    """
    args = sys.argv[1:]

    # 获取默认 workspace（从环境变量或 ~/Desktop）
    default_workspace = get_default_workspace()

    if not args:
        print("用法:")
        print("  推荐方式（环境变量）: WORKSPACE=~/Desktop python3 generate_excel.py --date 20260328")
        print("  新方式: python3 generate_excel.py --workspace ~/Desktop --date 20260328")
        print("  旧方式: python3 generate_excel.py ~/Desktop/20260328选品 20260328")
        print("")
        print("例如:")
        print(f"  python3 generate_excel.py --date 20260328")
        print(f"    （自动使用 workspace: {default_workspace}）")
        print("  WORKSPACE=/data/products python3 generate_excel.py --date 20260328")
        print("  python3 generate_excel.py --workspace ~/Desktop --date 20260328")
        sys.exit(1)

    # 判断调用方式
    if args[0].startswith('--'):
        # 新方式：解析 flags
        workspace = None
        date_str = None
        custom_json = None

        i = 0
        while i < len(args):
            if args[i] == '--workspace' and i + 1 < len(args):
                workspace = args[i + 1]
                i += 2
            elif args[i] == '--date' and i + 1 < len(args):
                date_str = args[i + 1]
                i += 2
            elif args[i] == '--custom' and i + 1 < len(args):
                custom_json = args[i + 1]
                i += 2
            else:
                i += 1

        if not date_str:
            print("❌ 缺少必要参数: --date")
            sys.exit(1)

        # workspace 未指定时使用默认值（来自环境变量或 ~/Desktop）
        workspace = workspace or default_workspace

        # 创建输出目录：{workspace}/{date}选品/
        workspace = os.path.expanduser(workspace)
        output_dir = str(Path(workspace) / f"{date_str}选品")
        os.makedirs(output_dir, exist_ok=True)
        return output_dir, date_str, custom_json
    else:
        # 旧方式：直接位置参数
        output_dir = args[0]
        date_str = args[1] if len(args) > 1 else ""
        custom_json = args[2] if len(args) > 2 else None

        if not date_str:
            print("❌ 旧方式缺少必要参数: <OUTPUT_DIR> <DS>")
            sys.exit(1)

        return output_dir, date_str, custom_json

OUTPUT_DIR, DS, CUSTOM_JSON = parse_args()
# 已排查产品.xlsx 位于选品目录的父目录（通常是桌面或workspace根目录）
REJECTED_FILE = str(Path(OUTPUT_DIR).parent / "已排查产品.xlsx")


# ===== Step 1: 解析数据 =====
def parse_products(output_dir):
    """自动发现目录下所有 data-get-第x页数据.json 并合并解析"""
    pattern = os.path.join(output_dir, "data-get-第*页数据.json")
    files = sorted(glob.glob(pattern))
    if not files:
        print(f"  ⚠️ 未找到数据文件（{pattern}）")
        sys.exit(1)
    print(f"  发现数据文件：{[os.path.basename(f) for f in files]}")

    all_products = []
    for filepath in files:
        with open(filepath, 'r', encoding='utf-8') as f:
            content = f.read()
        # 尝试 JSON 格式解析（结构化数据）
        if content.strip().startswith('[') or content.strip().startswith('{'):
            try:
                json_data = json.loads(content)
                if isinstance(json_data, dict) and 'raw_mcp_response' in json_data:
                    raise ValueError('placeholder format')
                if isinstance(json_data, list):
                    for item in json_data:
                        p = {
                            '序号': str(len(all_products) + 1),
                            '分区日期': item.get('ds', DS),
                            'SKU_ID': item.get('SKU_ID', ''),
                            'SPU_ID': item.get('SPU_ID', item.get('SKU_ID', '')),
                            '英文标题': item.get('title', ''),
                            '商品链接': item.get('url', ''),
                            '主图链接': item.get('image', ''),
                            '一级类目': item.get('category', '').split('->')[0] if '->' in item.get('category', '') else item.get('category', ''),
                            '类目路径': item.get('category', ''),
                            '一级类目ID': '',
                            '类目ID路径': '',
                            '售价': str(item.get('price', '')),
                            '货币': item.get('currency', '€'),
                            '评分次数': str(item.get('reviews', '')),
                            '评分星级': str(item.get('rating', '')),
                            'SPU月销量': str(item.get('spu_sales_30d', '')),
                            'SKU月销量': str(item.get('sku_sales_30d', '')),
                            '变体数量': str(item.get('variants', '')),
                            '卖家名称': item.get('seller', ''),
                            '体积': item.get('volume', ''),
                            '重量': item.get('weight', ''),
                            '采集时间': '',
                            '五点描述': '',
                        }
                        all_products.append(p)
                    print(f"  JSON格式解析完成，共 {len(json_data)} 条")
                    continue
            except (json.JSONDecodeError, ValueError):
                pass  # 不是有效JSON，走文本解析
        # 原始 MCP 文本格式解析
        blocks = re.split(r'————————————————————————————————', content)
        for block in blocks:
            block = block.strip()
            if not block or '商品序号：' not in block:
                continue
            p = {}
            def ext(pattern, default=''):
                m = re.search(pattern, block)
                return m.group(1).strip() if m else default
            p['序号']     = ext(r'商品序号：(\d+)')
            p['分区日期'] = ext(r'分区日期：(\d+)')
            p['SKU_ID']   = ext(r'SKU ID：(.+)')
            p['SPU_ID']   = ext(r'SPU ID：(.+)')
            p['英文标题'] = ext(r'商品标题：(.+)')
            p['商品链接'] = ext(r'商品链接：(https?://\S+)')
            p['主图链接'] = ext(r'主图链接：(https?://\S+)')
            p['一级类目'] = ext(r'一级类目名称：(.+)')
            p['类目路径'] = ext(r'类目名称路径：(.+)')
            p['一级类目ID'] = ext(r'一级类目ID：(.+)')
            # 类目ID路径：原始格式是 -> 分隔，转换为 : 分隔（MCP调用需要）
            node_id_path = ext(r'类目ID路径：(.+)')
            p['类目ID路径'] = node_id_path.replace('->', ':') if node_id_path else ''
            p['售价']     = ext(r'售价：(.+)')
            p['货币']     = ext(r'货币单位：(.+)')
            p['评分次数'] = ext(r'评分次数：(.+)')
            p['评分星级'] = ext(r'评分星级：(.+)')
            p['SPU月销量']= ext(r'近30天SPU销量：(.+)')
            p['SKU月销量']= ext(r'近30天SKU销量：(.+)')
            p['变体数量'] = ext(r'变体数量：(.+)')
            p['卖家名称'] = ext(r'卖家名称：(.+)')
            p['体积']     = ext(r'体积：(.+)')
            p['重量']     = ext(r'重量：(.+)')
            p['采集时间'] = ext(r'采集时间：(.+)')
            # 五点描述
            fm = re.search(r'五点描述：(\[.*?\])', block, re.DOTALL)
            if fm:
                try:
                    fl = json.loads(fm.group(1))
                    p['五点描述'] = ' | '.join(fl[:3])
                except Exception:
                    p['五点描述'] = ''
            else:
                p['五点描述'] = ''
            # 详细描述
            dm = re.search(r'详细描述：(.+?)(?=\n采集时间|$)', block, re.DOTALL)
            p['详细描述'] = dm.group(1).strip() if dm and dm.group(1).strip() != 'null' else ''
            if p['SKU_ID']:
                all_products.append(p)

    return all_products


# ===== Step 2: 读取 AI 对话预处理结果 =====
def load_ai_results(output_dir):
    ai_file = os.path.join(output_dir, 'ai_results.json')
    if os.path.exists(ai_file):
        with open(ai_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        print(f"  已读取 ai_results.json，共 {len(data)} 条")
        # 转换为字典格式，以SKU_ID为key
        return {item['SKU_ID']: item for item in data}
    print("  ⚠️ ai_results.json 不存在，使用规则兜底（中文标题取英文前4词，复筛全部通过）")
    return {}


def apply_ai_results(products, output_dir):
    ai_map = load_ai_results(output_dir)
    fallback_count = 0
    for p in products:
        result = ai_map.get(p['SKU_ID'], {})
        cn = result.get('中文标题', '')
        if cn and re.search(r'[\u4e00-\u9fff]', cn):
            p['中文标题'] = cn[:18]
        else:
            words = p['英文标题'].split()
            p['中文标题'] = ' '.join(words[:4])[:20]
            fallback_count += 1
        p['复筛结论']    = result.get('复筛结论', '通过')
        p['复筛备注']    = result.get('剔除原因', '')
        p['季节性产品']  = result.get('季节性产品', '否')
        p['季节性关键词']= result.get('节日名称', '')
        p['节日月份']    = result.get('节日月份', '')
    if fallback_count:
        print(f"  规则兜底：{fallback_count} 条（未在 ai_results.json 中找到）")
    print(f"  ✅ AI 预处理结果已填充，共 {len(products)} 条")
    return ai_map  # 返回ai_map用于条数检查


# ===== 工具函数 =====
def make_header_cell(cell, text, bg='1F4E79', fg='FFFFFF', bold=True, center=True):
    cell.value = text
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.font = Font(bold=bold, color=fg, name="Arial", size=10)
    if center:
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


# ===== Step 1b: 从 merged_products.json 或自定义JSON读取（推荐模式） =====
def load_merged_products(output_dir, custom_json=None):
    """读取 merged_products.json 或自定义JSON文件（AI 在上下文中一次性完成提取+预处理的结果）"""
    merged_file = os.path.join(output_dir, custom_json) if custom_json else os.path.join(output_dir, 'merged_products.json')
    if not os.path.exists(merged_file):
        print(f"  ⚠️ 未找到数据文件：{merged_file}")
        return None
    with open(merged_file, 'r', encoding='utf-8') as f:
        data = json.load(f)
    if not isinstance(data, list):
        return None
    # 确保每条记录有完整字段
    required_keys = {'SKU_ID', '英文标题', '售价'}
    valid = [p for p in data if required_keys.issubset(p.keys())]
    if len(valid) == 0:
        return None
    print(f"  ✅ 从 merged_products.json 读取，共 {len(valid)} 条")
    return valid


# ===== 主流程 =====
def main():
    print(f"=== 选品 Excel 生成 | {DS} | {OUTPUT_DIR} ===\n")

    # 优先模式：merged_products.json 或自定义JSON文件（AI 上下文中一次完成提取+预处理）
    products = load_merged_products(OUTPUT_DIR, CUSTOM_JSON)

    if products is not None:
        # merged 模式：字段已经完整，直接填充默认值即可
        ai_count = len(products)
        products_count = len(products)
        print(f"  模式：merged（推荐，无原始数据文件依赖）\n")
        # 补齐缺失的可选字段
        for p in products:
            p.setdefault('序号', str(products.index(p) + 1))
            p.setdefault('分区日期', DS)
            p.setdefault('SPU_ID', p.get('SKU_ID', ''))
            p.setdefault('商品链接', p.get('url', ''))
            p.setdefault('主图链接', p.get('image', ''))
            p.setdefault('一级类目', (p.get('category', '') or '').split('->')[0] if '->' in (p.get('category', '') or '') else (p.get('category', '') or ''))
            p.setdefault('类目路径', p.get('category', ''))
            p.setdefault('一级类目ID', '')
            p.setdefault('类目ID路径', '')
            p.setdefault('货币', p.get('currency', '€'))
            p.setdefault('评分次数', str(p.get('reviews', '')))
            p.setdefault('评分星级', str(p.get('rating', '')))
            p.setdefault('SPU月销量', str(p.get('spu_sales_30d', '')))
            p.setdefault('SKU月销量', str(p.get('sku_sales_30d', '')))
            p.setdefault('变体数量', str(p.get('variants', '')))
            p.setdefault('卖家名称', p.get('seller', ''))
            p.setdefault('体积', p.get('volume', ''))
            p.setdefault('重量', p.get('weight', ''))
            p.setdefault('采集时间', p.get('collect_time', ''))
            p.setdefault('五点描述', '')
            # AI 预处理字段（merged 中已有，补默认值防缺失）
            cn = p.get('中文标题', '')
            if not cn or not re.search(r'[\u4e00-\u9fff]', cn):
                words = p['英文标题'].split()
                p['中文标题'] = ' '.join(words[:4])[:20]
            p.setdefault('复筛结论', '通过')
            p.setdefault('复筛备注', p.get('剔除原因', ''))
            p.setdefault('季节性产品', p.get('seasonal', '否'))
            p.setdefault('季节性关键词', p.get('节日名称', ''))
            p.setdefault('节日月份', p.get('节日月份', ''))
    else:
        # 兼容模式：原始双文件流程（data-get + ai_results）
        print("Step 1: 解析原始数据文件")
        products = parse_products(OUTPUT_DIR)
        products_count = len(products)
        print(f"  共解析 {products_count} 条\n")

        print("Step 2: 读取 AI 预处理结果（中文标题 + 复筛 + 季节性）")
        ai_map = load_ai_results(OUTPUT_DIR)
        ai_count = len(ai_map)
        print(f"  ai_results.json 条数: {ai_count}")

        if ai_count != products_count:
            print(f"\n⚠️  警告：数据条数不一致！")
            print(f"  原始数据: {products_count} 条")
            print(f"  AI结果: {ai_count} 条")
            print(f"  差异: {abs(ai_count - products_count)} 条")

        apply_ai_results(products, OUTPUT_DIR)
        print()

    print("Step 3: 分类（通过/剔除）")
    passed_all, rejected = [], []
    for p in products:
        (passed_all if p['复筛结论'] == '通过' else rejected).append(p)
    
    # 按 SPU_ID 去重，保留每个SPU的第一条
    seen_spu = set()
    passed = []
    for p in passed_all:
        spu = p.get('SPU_ID', '')
        if spu and spu not in seen_spu:
            seen_spu.add(spu)
            passed.append(p)
    
    print(f"  通过: {len(passed)} (去重前: {len(passed_all)})  剔除: {len(rejected)}\n")

    # ---------- 选品表 ----------
    print("Step 4: 生成选品.xlsx")
    wb1 = Workbook()
    ws1 = wb1.active
    ws1.title = "选品表"
    headers1 = ['序号','中文标题','英文标题','SKU_ID','SPU_ID','售价','货币',
                '评分星级','评分次数','SPU月销量','SKU月销量','变体数量',
                '一级类目','类目路径','卖家名称','重量','体积',
                '商品链接','主图链接','分区日期','采集时间']
    for c, h in enumerate(headers1, 1):
        make_header_cell(ws1.cell(1, c), h)
    ws1.row_dimensions[1].height = 30
    for ri, p in enumerate(products, 2):
        row = [p['序号'],p['中文标题'],p['英文标题'],p['SKU_ID'],p['SPU_ID'],
               p['售价'],p['货币'],p['评分星级'],p['评分次数'],
               p['SPU月销量'],p['SKU月销量'],p['变体数量'],
               p['一级类目'],p['类目路径'],p['卖家名称'],p['重量'],p['体积'],
               p['商品链接'],p['主图链接'],p['分区日期'],p['采集时间']]
        for c, v in enumerate(row, 1):
            cell = ws1.cell(ri, c, v)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical='center')
        if ri % 2 == 0:
            for c in range(1, len(headers1)+1):
                ws1.cell(ri, c).fill = PatternFill("solid", fgColor="EBF3FB")
    for i, w in enumerate([5,18,50,14,14,8,5,8,8,10,10,8,15,45,15,10,15,50,50,10,18], 1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = 'A2'
    out1 = os.path.join(OUTPUT_DIR, f"{DS}选品.xlsx")
    wb1.save(out1)
    print(f"  ✅ {out1}\n")

    # ---------- 复筛表 ----------
    print("Step 5: 生成复筛.xlsx")
    wb2 = Workbook()
    ws2 = wb2.active
    ws2.title = "复筛通过"
    headers2 = ['序号','中文标题','英文标题','SKU_ID','SPU_ID','售价','货币',
                '评分星级','评分次数','SPU月销量','SKU月销量','变体数量',
                '一级类目','类目路径','一级类目ID','类目ID路径','卖家名称','重量','体积',
                '复筛结论','季节性产品','季节性关键词','节日月份','复筛备注','数据分区日期',
                '商品链接','主图链接','采集时间']
    for c, h in enumerate(headers2, 1):
        make_header_cell(ws2.cell(1, c), h)
    ws2.row_dimensions[1].height = 30
    for ri, p in enumerate(passed, 2):
        row = [p['序号'],p['中文标题'],p['英文标题'],p['SKU_ID'],p['SPU_ID'],
               p['售价'],p['货币'],p['评分星级'],p['评分次数'],
               p['SPU月销量'],p['SKU月销量'],p['变体数量'],
               p['一级类目'],p['类目路径'],p.get('一级类目ID',''),p.get('类目ID路径',''),
               p['卖家名称'],p['重量'],p['体积'],
               '通过',p['季节性产品'],p['季节性关键词'],p['节日月份'],p['复筛备注'],DS,
               p['商品链接'],p['主图链接'],p['采集时间']]
        for c, v in enumerate(row, 1):
            cell = ws2.cell(ri, c, v)
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical='center')
        if ri % 2 == 0:
            for c in range(1, len(headers2)+1):
                ws2.cell(ri, c).fill = PatternFill("solid", fgColor="E8F5E9")
        if p.get('季节性产品') == '是':
            ws2.cell(ri, 19).font = Font(name="Arial", size=9, color="C00000", bold=True)
    # 剔除 sheet
    ws3 = wb2.create_sheet("被剔除产品")
    headers3 = ['序号','中文标题','英文标题','SKU_ID','类目路径','售价','复筛结论','剔除原因','数据分区日期']
    for c, h in enumerate(headers3, 1):
        make_header_cell(ws3.cell(1, c), h, bg='8B0000')
    for ri, p in enumerate(rejected, 2):
        row = [p['序号'],p['中文标题'],p['英文标题'],p['SKU_ID'],p['类目路径'],
               p['售价'],'剔除',p['复筛备注'],DS]
        for c, v in enumerate(row, 1):
            ws3.cell(ri, c, v).font = Font(name="Arial", size=9)
    for i, w in enumerate([5,18,50,14,14,8,5,8,8,10,10,8,15,45,12,35,15,10,15,8,8,15,8,30,10,50,50,18], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = 'A2'
    out2 = os.path.join(OUTPUT_DIR, f"{DS}选品复筛.xlsx")
    wb2.save(out2)
    print(f"  ✅ {out2}\n")

    # ---------- 已排查产品 ----------
    print("Step 6: 追加已排查产品.xlsx")
    if os.path.exists(REJECTED_FILE):
        wb_rej = load_workbook(REJECTED_FILE)
        ws_rej = wb_rej.active
        next_row = ws_rej.max_row + 1
    else:
        wb_rej = Workbook()
        ws_rej = wb_rej.active
        ws_rej.title = "已排查产品"
        for c, h in enumerate(['SKU_ID','中文标题','英文标题','类目路径','剔除原因','数据分区日期','追加时间'], 1):
            make_header_cell(ws_rej.cell(1, c), h, bg='7B2D00')
        next_row = 2
    now_str = datetime.now().strftime('%Y-%m-%d %H:%M')
    for p in rejected:
        for c, v in enumerate([p['SKU_ID'],p['中文标题'],p['英文标题'],
                                p['类目路径'],p['复筛备注'],DS,now_str], 1):
            ws_rej.cell(next_row, c, v).font = Font(name="Arial", size=9)
        next_row += 1
    wb_rej.save(REJECTED_FILE)
    print(f"  ✅ 已追加 {len(rejected)} 条 → {REJECTED_FILE}\n")

    print("=== 完成 ===")
    print(f"  总条数: {len(products)}  通过: {len(passed)}  剔除: {len(rejected)}")
    print(f"  AI结果条数: {ai_count}  原始数据条数: {products_count}")
    if ai_count == products_count:
        print(f"  ✅ 数据条数一致性检查通过")
    else:
        print(f"  ⚠️  数据条数不一致，部分产品使用了规则兜底")
    seasonal_cnt = sum(1 for p in passed if p.get('季节性产品') == '是')
    print(f"  季节性产品（通过中）: {seasonal_cnt}")
    print(f"  输出目录: {OUTPUT_DIR}")


if __name__ == '__main__':
    main()
