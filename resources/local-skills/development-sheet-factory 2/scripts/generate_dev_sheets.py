#!/usr/bin/env python3
"""
批量生成开发表格 —— development-sheet-factory 核心脚本
=======================================================
将选品复筛表批量转换为 SPU 开发表格，可选同时填写竞品数据。

用法:
  # 仅生成基础字段（不含竞品）
  python3 generate_dev_sheets.py <复筛表路径> <日期YYYYMMDD>

  # 一次性生成全部字段（基础 + 竞品）
  python3 generate_dev_sheets.py <复筛表路径> <日期YYYYMMDD> --competitor <竞品数据.json>

竞品数据 JSON 格式（由 AI agent 调用 sellersprite product_research + review 后汇总生成）:
  {
    "B0XXXXXX": {
      "product_research": {"items": [...]},  // 5个竞品（变体<6，价格±30%）
      "reviews": {"B0AAAAAA": [...], "B0BBBBBB": [...]}  // 评分最低3个竞品的评论
    },
    ...
  }
  键为 SKU_ID（原始 ASIN）。

执行流程:
  1. 读取复筛表
  2. （若提供 --competitor）加载竞品数据 JSON
  3. 对每个 SPU，复制模板 → 打开一次 → 写基础字段 + 竞品字段 → 保存
  4. 写入 _READY.json
"""

import sys
import os
import json
import shutil
import re
import glob
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.utils.units import pixels_to_EMU
    import urllib.request
    from PIL import Image as PILImage
except ImportError as e:
    print(f"❌ 缺少依赖: {e}\n请运行: pip3 install openpyxl Pillow")
    sys.exit(1)

# ── 常量 ──────────────────────────────────────────────────────────────────────

SKILL_DIR     = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_PATH = os.path.join(SKILL_DIR, "assets", "开发模版.xlsx")
_PRICE_RE     = re.compile(r'[\d.]+')   # 预编译，避免重复编译

# ── 工具函数 ──────────────────────────────────────────────────────────────────

def safe_filename(name: str) -> str:
    """清理文件名中的非法字符，限制长度"""
    name = re.sub(r'[\\/:*?"<>|]', '_', name)
    return name[:50]


def parse_price(price_str):
    """从 '€15.00' / '15.99' 等格式提取浮点数"""
    if not price_str:
        return None
    m = _PRICE_RE.search(str(price_str))
    return float(m.group()) if m else None


def download_image(url: str, tmp_path: str) -> bool:
    """下载图片到临时文件，返回是否成功"""
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=20) as resp:
            with open(tmp_path, 'wb') as f:
                f.write(resp.read())
        return os.path.getsize(tmp_path) > 100
    except Exception as e:
        print(f"    ⚠️  图片下载失败: {e}")
        return False


# ── 竞品数据解析 ──────────────────────────────────────────────────────────────

def extract_negative_keywords(reviews: list) -> str:
    """
    从评论列表中提取差评关键词
    分析1-3星评论，提取高频问题词
    """
    if not reviews:
        return "暂无评论数据"
    
    # 筛选1-3星差评
    negative_reviews = [r for r in reviews if r.get('rating', 5) <= 3]
    if not negative_reviews:
        return "差评较少，整体口碑良好"
    
    # 定义关键词映射（中英文）
    keyword_patterns = {
        '质量': ['质量', 'quality', '劣质', '差', 'bad quality', 'poor quality'],
        '尺寸': ['尺寸', '大小', 'size', '太小', '太大', 'small', 'large', 'tight'],
        '材质': ['材质', '材料', 'material', '塑料', '薄', 'cheap material'],
        '功能': ['功能', '不工作', 'broken', 'not work', 'defective', 'issue'],
        '包装': ['包装', '破损', 'package', 'damaged', 'broken box'],
        '物流': ['物流', '慢', 'shipping', 'delivery', 'late'],
        '描述不符': ['不符', '描述', 'different', 'not as described', 'misleading'],
    }
    
    # 统计关键词出现次数
    keyword_counts = {}
    for review in negative_reviews:
        title = str(review.get('title', '')).lower()
        content = str(review.get('content', '')).lower()
        text = title + ' ' + content
        
        for category, patterns in keyword_patterns.items():
            for pattern in patterns:
                if pattern.lower() in text:
                    keyword_counts[category] = keyword_counts.get(category, 0) + 1
                    break
    
    if not keyword_counts:
        return f"有{len(negative_reviews)}条差评，未提取到明确问题类型"
    
    # 按出现次数排序，取前3个
    sorted_keywords = sorted(keyword_counts.items(), key=lambda x: x[1], reverse=True)[:3]
    keywords_str = '、'.join([k for k, v in sorted_keywords])
    
    return f"主要问题：{keywords_str}（基于{len(negative_reviews)}条差评）"


def parse_competitor(data: dict) -> dict:
    """
    从 traffic_listing 竞品数据结构中提取 D2/E2/I2/S2/T2 所需数据。

    参数:
        data: {
            "traffic_listing": {"items": [...], "data": {...}}
        }

    返回:
        {
            "link1":          D2 竞品链接1
            "link2":          E2 竞品链接2
            "price_range":    I2 竞品定价区间
            "avg_rating":     S2 市场类目评分
            "rating_summary": T2 市场差评说明
        }
    """
    if not data:
        return {}
    
    # 兼容旧格式（直接是items列表）
    if isinstance(data, list):
        data = {'traffic_listing': {'items': data, 'data': {}}}
    
    items = data.get('traffic_listing', {}).get('items', [])
    if not items:
        return {}

    # 取前3个有效竞品
    valid = [it for it in items if it.get('asin')][:3]
    if not valid:
        return {}

    result = {}
    base_url = "https://www.amazon.de/dp/"

    # D2 / E2
    if len(valid) >= 1:
        result['link1'] = base_url + valid[0]['asin']
    if len(valid) >= 2:
        result['link2'] = base_url + valid[1]['asin']

    # I2：定价区间（基于所有竞品）
    all_items = [it for it in items if it.get('price')]
    if all_items:
        prices = [it.get('price') for it in all_items]
        lo, hi = min(prices), max(prices)
        if abs(hi - lo) < 0.01:
            result['price_range'] = f"€{lo:.2f}"
        else:
            result['price_range'] = f"€{lo:.2f} - €{hi:.2f}"

    # S2：市场评分（基于所有竞品的平均值）
    all_ratings = [it.get('rating') for it in items if it.get('rating')]
    if all_ratings:
        avg = round(sum(all_ratings) / len(all_ratings), 1)
        result['avg_rating'] = f"{avg}/5.0"

        # T2：评分描述
        low_count = sum(1 for r in all_ratings if r < 4.0)
        n = len(all_ratings)
        if low_count == 0:
            result['rating_summary'] = f"基于{n}个竞品，均分{avg}，整体口碑良好"
        else:
            result['rating_summary'] = f"基于{n}个竞品，均分{avg}，{low_count}个评分偏低"

    return result


# ── 图片并发预下载 ────────────────────────────────────────────────────────────

def prefetch_images(rows: list, tmp_dir: str = "/tmp") -> dict:
    """
    并发下载所有行的主图，返回 {index: tmp_path or None}。
    index 对应 rows 的下标（从1开始）。
    """
    def _fetch(args):
        idx, url = args
        if not url:
            return idx, None
        tmp_path = os.path.join(tmp_dir, f"_dev_img_{idx}.jpg")
        ok = download_image(url, tmp_path)
        return idx, tmp_path if ok else None

    tasks = [(i, row.get('主图链接', '')) for i, row in enumerate(rows, 1)]
    results = {}

    print(f"  🖼️  并发下载 {len(tasks)} 张主图...")
    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = {pool.submit(_fetch, t): t[0] for t in tasks}
        for fut in as_completed(futures):
            idx, path = fut.result()
            results[idx] = path

    ok_count = sum(1 for v in results.values() if v)
    print(f"  🖼️  图片下载完成: {ok_count}/{len(tasks)} 成功")
    return results


# ── 单文件生成（基础字段 + 竞品字段，只打开一次）────────────────────────────

def _calc_f2_size(ws) -> tuple:
    """
    计算 F2:H2 合并单元格的像素尺寸，用于自适应图片缩放。
    - F、G、H三列合并，所以宽度是三列之和
    - 列宽（字符数）→ 像素：×7（Excel 默认字体 "0" 字符宽度约 7px）
    - 行高（磅）→ 像素：×96/72（96 DPI 下 72 磅 = 1 英寸）
    """
    # F、G、H三列宽度之和
    col_w_f = ws.column_dimensions['F'].width or 13
    col_w_g = ws.column_dimensions['G'].width or 13
    col_w_h = ws.column_dimensions['H'].width or 13
    total_col_w = col_w_f + col_w_g + col_w_h
    
    row_h = ws.row_dimensions[2].height or 147  # 磅 points
    PX_PER_CHAR = 7
    PX_PER_PT   = 96 / 72
    cell_w = total_col_w * PX_PER_CHAR
    cell_h = row_h * PX_PER_PT
    return cell_w, cell_h


def _fit_image(img_path: str, max_w: float, max_h: float) -> tuple:
    """
    计算图片在 max_w × max_h 限制内的缩放尺寸，保持宽高比。
    使用 contain 模式，确保图片完整显示在单元格内，不溢出。
    返回 (target_w, target_h)。
    """
    with PILImage.open(img_path) as im:
        orig_w, orig_h = im.size
    # 使用 contain 模式：取较小的缩放比例，确保图片不超出单元格
    ratio = min(max_w / orig_w, max_h / orig_h)
    return orig_w * ratio, orig_h * ratio


def generate_sheet(row_data: dict, output_dir: str, index: int,
                   img_path=None,
                   competitor=None) -> str:
    """
    为单个 SPU 生成开发表格，返回输出文件路径。
    - img_path:   预下载的图片临时路径（None 则退回写 URL）
    - competitor: parse_competitor() 返回的竞品数据（None 则不填竞品字段）
    """
    sku_id        = row_data.get('SKU_ID') or row_data.get('SPU_ID', '')
    spu_id        = row_data.get('SPU_ID') or sku_id
    cn_title      = row_data.get('中文标题', '')
    link          = row_data.get('商品链接', '')
    image_url     = row_data.get('主图链接', '')
    price_str     = row_data.get('售价', '')
    is_seasonal   = row_data.get('季节性产品', '否')
    seasonal_month= row_data.get('节日月份', '')

    price    = parse_price(price_str)
    fname    = safe_filename(f"{spu_id}+{cn_title}") + ".xlsx"
    out_path = os.path.join(output_dir, fname)

    # 直接从模板文件加载，内存修改后保存到目标路径（省去中间 copy2 + 二次读）
    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # ── 基础字段 ──
    ws['C2'] = link
    ws['O2'] = is_seasonal
    ws['P2'] = seasonal_month or ''
    ws['B7'] = cn_title
    if price is not None:
        ws['I7'] = price
        ws['J7'] = price

    # F2:H2：主图，填满合并单元格
    if img_path:
        try:
            cell_w, cell_h = _calc_f2_size(ws)
            target_w, target_h = _fit_image(img_path, cell_w, cell_h)
            img = XLImage(img_path)
            img.width  = target_w
            img.height = target_h
            
            # 使用 OneCellAnchor 让图片锚定到 F2:H2 合并单元格
            from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU
            
            # 创建锚点：从 F2 单元格开始（F=5 0-indexed, row=1 0-indexed for row 2）
            marker = AnchorMarker(col=5, row=1, colOff=0, rowOff=0)
            
            # 创建尺寸对象
            size = XDRPositiveSize2D(cx=pixels_to_EMU(target_w), cy=pixels_to_EMU(target_h))
            
            # 创建 OneCellAnchor
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            
            ws.add_image(img)
        except Exception as e:
            print(f"    ⚠️  图片插入失败: {e}")
            if image_url:
                ws['F2'] = image_url
    elif image_url:
        ws['F2'] = image_url   # 下载失败时退回写 URL

    # ── 竞品字段（若提供）──
    if competitor:
        cell_map = {
            'link1':          'D2',
            'link2':          'E2',
            'price_range':    'I2',
            'avg_rating':     'S2',
            'rating_summary': 'T2',
        }
        for key, cell in cell_map.items():
            val = competitor.get(key, '')
            if val:
                ws[cell] = val

    wb.save(out_path)
    return out_path


# ── 主流程 ────────────────────────────────────────────────────────────────────

def main():
    # ── 参数解析 ──
    args = sys.argv[1:]

    # 获取默认 workspace（从环境变量或 ~/Desktop）
    from pathlib import Path
    default_workspace = os.environ.get('WORKSPACE', os.path.expanduser('~/Desktop'))

    if len(args) < 1:
        print("用法:")
        print("  旧方式: python3 generate_dev_sheets.py <复筛表> <YYYYMMDD>")
        print("  新方式: python3 generate_dev_sheets.py --workspace <WORKSPACE> --date <YYYYMMDD>")
        print("  推荐方式（环境变量）: WORKSPACE=~/Desktop python3 generate_dev_sheets.py --date 20260328")
        print("")
        print("可选参数:")
        print("  --competitor <JSON>  竞品数据文件")
        print("  --limit <N>          只生成前N个（测试模式）")
        print("  --spu <SKU1,SKU2>    指定SPU列表（增量生成）")
        print("")
        print("环境变量:")
        print(f"  WORKSPACE            默认workspace (当前: {default_workspace})")
        print("")
        print("例如:")
        print("  python3 generate_dev_sheets.py ~/Desktop/20260328选品/20260328选品复筛.xlsx 20260328")
        print("  python3 generate_dev_sheets.py --workspace ~/Desktop --date 20260328")
        print("  python3 generate_dev_sheets.py --date 20260328  # 使用 ~/Desktop")
        print("  python3 generate_dev_sheets.py --date 20260328 --competitor /path/to/competitor.json")
        sys.exit(1)

    # 判断调用方式
    if args[0].startswith('--'):
        # 新方式：从 workspace 和 date 推导复筛表路径
        workspace = default_workspace
        date_str = None
        competitor_json = None
        limit_count = None
        filter_spus = None

        i = 0
        while i < len(args):
            if args[i] == '--workspace' and i + 1 < len(args):
                workspace = args[i + 1]
                i += 2
            elif args[i] == '--date' and i + 1 < len(args):
                date_str = args[i + 1]
                i += 2
            elif args[i] == '--competitor' and i + 1 < len(args):
                competitor_json = args[i + 1]
                i += 2
            elif args[i] == '--limit' and i + 1 < len(args):
                try:
                    limit_count = int(args[i + 1])
                    print(f"🧪 测试模式: 只生成前 {limit_count} 个开发表格")
                except ValueError:
                    print("❌ --limit 参数需要整数")
                    sys.exit(1)
                i += 2
            elif args[i] == '--spu' and i + 1 < len(args):
                spu_str = args[i + 1]
                filter_spus = set(s.strip() for s in spu_str.split(',') if s.strip())
                print(f"🎯 增量模式: 只生成指定SPU {filter_spus}")
                i += 2
            else:
                i += 1

        if not date_str:
            print("❌ 新方式缺少必要参数: --date 是必需的")
            sys.exit(1)

        # 自动推导复筛表路径：{workspace}/{date}选品/{date}选品复筛.xlsx
        source_xlsx = str(Path(workspace) / f"{date_str}选品" / f"{date_str}选品复筛.xlsx")
    else:
        # 旧方式：直接位置参数
        source_xlsx = args[0]

        if len(args) < 2:
            print("❌ 旧方式缺少必要参数: 需要 <复筛表> 和 <YYYYMMDD>")
            sys.exit(1)

        date_str = args[1]
        competitor_json = None
        limit_count = None
        filter_spus = None

        for i, arg in enumerate(args[2:], 2):
            if arg == '--competitor' and i + 1 < len(args):
                competitor_json = args[i + 1]
            if arg == '--limit' and i + 1 < len(args):
                try:
                    limit_count = int(args[i + 1])
                    print(f"🧪 测试模式: 只生成前 {limit_count} 个开发表格")
                except ValueError:
                    print("❌ --limit 参数需要整数")
                    sys.exit(1)
            if arg == '--spu' and i + 1 < len(args):
                spu_str = args[i + 1]
                filter_spus = set(s.strip() for s in spu_str.split(',') if s.strip())
                print(f"🎯 增量模式: 只生成指定SPU {filter_spus}")

    # ── 前置检查 ──
    if not os.path.exists(source_xlsx):
        print(f"❌ 找不到复筛表: {source_xlsx}")
        sys.exit(1)
    if not os.path.exists(TEMPLATE_PATH):
        print(f"❌ 找不到模板: {TEMPLATE_PATH}")
        sys.exit(1)

    # ── 读取复筛表 ──
    wb_src  = openpyxl.load_workbook(source_xlsx)
    ws_src  = wb_src.active
    headers = [cell.value for cell in ws_src[1]]
    rows    = []
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        if any(row):
            rows.append(dict(zip(headers, row)))
    print(f"📋 复筛表: {len(rows)} 个 SPU")

    # 应用 limit 限制（用于测试）
    if limit_count is not None and limit_count > 0:
        rows = rows[:limit_count]
        print(f"🧪 已限制为只处理前 {len(rows)} 个")

    # 应用 SPU 过滤（用于增量生成）
    if filter_spus:
        original_count = len(rows)
        rows = [r for r in rows if (r.get('SPU_ID') or r.get('SKU_ID', '')) in filter_spus]
        print(f"🎯 已过滤到 {len(rows)} 个指定SPU（原{original_count}个）")

    # ── 加载竞品数据（可选）──
    all_competitor = {}
    if competitor_json:
        if not os.path.exists(competitor_json):
            print(f"❌ 找不到竞品数据文件: {competitor_json}")
            sys.exit(1)
        with open(competitor_json, encoding='utf-8') as f:
            raw = json.load(f)
        # 预解析所有竞品数据（兼容新旧格式）
        for sku_id, data in raw.items():
            # traffic_listing格式: {"traffic_listing": {"items": [...], "data": {...}}}
            # 旧格式: {"items": [...]}
            if isinstance(data, dict) and 'traffic_listing' in data:
                parsed = parse_competitor(data)
            else:
                # 旧格式兼容
                items = data.get('items', []) if isinstance(data, dict) else data
                parsed = parse_competitor({'traffic_listing': {'items': items, 'data': {}}})
            if parsed:
                all_competitor[sku_id] = parsed
        print(f"📦 竞品数据: {len(all_competitor)}/{len(raw)} 个 SPU 有有效竞品")

    # ── 创建输出目录 ──
    output_dir = os.path.join(os.path.dirname(source_xlsx), f"{date_str}开发表格")
    os.makedirs(output_dir, exist_ok=True)
    print(f"📁 输出目录: {output_dir}\n")

    # ── 并发预下载所有图片 ──
    img_map = prefetch_images(rows)

    # ── 批量生成（每个文件只打开一次）──
    print()
    generated, failed = [], []
    for i, row in enumerate(rows, 1):
        spu   = row.get('SPU_ID') or row.get('SKU_ID', '?')
        sku   = row.get('SKU_ID') or spu
        title = row.get('中文标题', '')
        print(f"  [{i:02d}/{len(rows)}] {spu}  {title[:20]}")

        img_path   = img_map.get(i)
        competitor = all_competitor.get(sku) or all_competitor.get(spu)

        try:
            path = generate_sheet(row, output_dir, i,
                                  img_path=img_path,
                                  competitor=competitor)
            generated.append({"sku": spu, "file": os.path.basename(path)})
            comp_tag = " +竞品" if competitor else ""
            print(f"    ✅ {os.path.basename(path)}{comp_tag}")
        except Exception as e:
            print(f"    ❌ 失败: {e}")
            failed.append({"sku": spu, "error": str(e)})

    # ── 写 _READY.json ──
    # comp_filled 统计：对每个生成成功的行，检查对应 SKU_ID 是否有竞品数据
    sku_set = set()
    for row in rows:
        spu = row.get('SPU_ID') or row.get('SKU_ID', '')
        sku = row.get('SKU_ID') or spu
        if all_competitor.get(sku) or all_competitor.get(spu):
            sku_set.add(spu)
    comp_filled = sum(1 for e in generated if e['sku'] in sku_set)
    
    # 更新竞品数据来源标识
    competitor_source = "sellersprite_traffic_listing" if competitor_json else ""
    ready = {
        "schema_version": "1.0",
        "date":           date_str,
        "workspace":      "product-selection",
        "status":         "ready" if not failed else "partial",
        "total_spu":      len(rows),
        "generated":      len(generated),
        "failed":         len(failed),
        "generated_at":   datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "template_path":  TEMPLATE_PATH,
        "competitor_source": competitor_source,
        "competitor_filled": comp_filled,
        "competitor_skipped": len(generated) - comp_filled,
        "files":          generated,
        "errors":         failed,
    }
    ready_path = os.path.join(output_dir, "_READY.json")
    with open(ready_path, 'w', encoding='utf-8') as f:
        json.dump(ready, f, ensure_ascii=False, indent=2)

    # ── 汇总 ──
    print(f"\n{'='*50}")
    print(f"✅ 完成: {len(generated)} 个文件生成成功，{len(failed)} 个失败")
    if competitor_json:
        print(f"📊 竞品数据: {comp_filled} 个已填写，{len(generated) - comp_filled} 个无数据")
    print(f"📂 输出目录: {output_dir}")
    if failed:
        print(f"⚠️  失败列表: {[x['sku'] for x in failed]}")


if __name__ == '__main__':
    main()
