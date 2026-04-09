#!/usr/bin/env python3
"""
竞品数据整合脚本 —— development-sheet-factory
==============================================
两种运行模式：

模式A（AI辅助，推荐）：
  AI 调用 MCP 获取竞品数据 → 传入 JSON → 脚本整合输出

模式B（MCP CLI，可选）：
  脚本直接调用 mcp CLI 获取数据（需 mcp CLI 在 PATH）

用法（模式A - 推荐）:
  python3 collect_competitor_data.py --mode integrate \
      --source <复筛表.xlsx> \
      --mcp-results <mcp_results.json> \
      --output <输出目录>

用法（模式B）:
  python3 collect_competitor_data.py --mode collect \
      --source <复筛表.xlsx> \
      --output <输出目录> \
      [--workers 5]

输入格式（mcp_results.json）:
  {
    "B0XXXXXX": {
      "traffic_listing": {
        "items": [...],    // 关联竞品列表
        "data": {...}      // 原始数据
      }
    },
    ...
  }

输出:
  <输出目录>/_competitor_data.json: 整合后的竞品数据
"""

import sys
import os
import json
import re
import argparse
import subprocess
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("❌ 缺少 openpyxl: pip3 install openpyxl")
    sys.exit(1)

# ── 常量 ──────────────────────────────────────────────────────────────────────

SKU_RE = re.compile(r'^B[A-Z0-9]{9}$')

# ── 工具函数 ──────────────────────────────────────────────────────────────────

def is_valid_asin(text: str) -> bool:
    return bool(SKU_RE.match(str(text).strip()))

def read_skus_from_xlsx(xlsx_path: str) -> tuple:
    """
    从复筛表读取 SKU 列表及相关字段
    
    Returns:
        (sku_list, price_map, node_id_map)
        - sku_list: SKU 列表
        - price_map: {sku: price}
        - node_id_map: {sku: nodeIdPath}
    """
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]

    # 查找列索引
    sku_col_idx = None
    price_col_idx = None
    node_id_col_idx = None
    
    for idx, h in enumerate(headers):
        if h:
            h_str = str(h).upper()
            if 'SKU' in h_str and 'ID' in h_str:
                sku_col_idx = idx
            elif h_str == 'PRICE' or h_str == '售价':
                price_col_idx = idx
            elif '类目ID' in h_str or h_str == 'NODEIDPATH':
                node_id_col_idx = idx

    if sku_col_idx is None:
        raise ValueError("未找到 SKU_ID 列")

    sku_list = []
    price_map = {}
    node_id_map = {}
    
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = row[sku_col_idx]
        if sku and is_valid_asin(str(sku)):
            sku_str = str(sku).strip()
            sku_list.append(sku_str)
            
            # 价格（处理 €14.99 或 14.99€ 格式）
            if price_col_idx is not None:
                price_val = row[price_col_idx]
                if price_val:
                    try:
                        # 移除货币符号和空格
                        price_str = str(price_val).replace('€', '').replace('$', '').replace('£', '').replace(',', '.').strip()
                        price_map[sku_str] = float(price_str)
                    except (ValueError, TypeError):
                        pass
            
            # 类目ID
            if node_id_col_idx is not None:
                node_id_val = row[node_id_col_idx]
                if node_id_val:
                    node_id_map[sku_str] = str(node_id_val).strip()

    return sku_list, price_map, node_id_map

def call_mcp_traffic_listing(asin: str) -> dict:
    """调用 sellersprite-mcp traffic_listing 获取关联竞品"""
    try:
        cmd = [
            "mcp", "call", "sellersprite-mcp", "traffic_listing",
            "--request", json.dumps({
                "marketplace": "DE",
                "asinList": [asin],
                "relations": ["similar"],
                "size": 10,
                "order": {"field": "total_units", "desc": True}
            })
        ]
        result = subprocess.run(
            cmd, capture_output=True, text=True, timeout=30
        )
        if result.returncode == 0:
            data = json.loads(result.stdout)
            items = data.get("data", {}).get("items", []) or data.get("items", [])
            return {"items": items, "data": data}
        else:
            # 打印错误日志便于调试
            print(f"    ⚠️ MCP调用失败: {result.stderr[:200]}")
            return {"items": [], "data": {}}
    except Exception as e:
        print(f"    ⚠️ MCP异常: {str(e)[:100]}")
        return {"items": [], "data": {}}


def collect_competitor_data(sku: str, price: float, node_id: str) -> dict:
    """
    收集单个SKU的竞品数据：
    调用 traffic_listing 获取关联竞品列表
    """
    result = {"traffic_listing": {"items": [], "data": {}}}
    
    # 获取关联竞品
    tl_data = call_mcp_traffic_listing(sku)
    items = tl_data.get("items", [])
    if items:
        result["traffic_listing"] = tl_data
    
    return result


def prefetch_all(sku_list: list, price_map: dict = None, node_id_map: dict = None, max_workers: int = 3) -> dict:
    """
    并发获取所有 SKU 的竞品数据
    
    Args:
        sku_list: SKU列表
        price_map: SKU到价格的映射 {sku: price}
        node_id_map: SKU到类目ID的映射 {sku: nodeIdPath}
    """
    results = {}
    total = len(sku_list)
    completed = 0
    
    price_map = price_map or {}
    node_id_map = node_id_map or {}

    print(f"📡 获取 {total} 个 SKU 的竞品数据（串行）...\n")

    for sku in sku_list:
        completed += 1
        try:
            results[sku] = collect_competitor_data(sku, price_map.get(sku, 20.0), node_id_map.get(sku, ""))
            has_data = results[sku]["traffic_listing"].get("items")
            status = "✅" if has_data else "⭕"
            item_count = len(results[sku]["traffic_listing"].get("items", []))
            print(f"  [{completed:02d}/{total}] {status} {sku} (items: {item_count})")
        except Exception as e:
            results[sku] = {"traffic_listing": {"items": [], "data": {}}}
            print(f"  [{completed:02d}/{total}] ❌ {sku}: {e}")

    return results

def integrate_data(sku_list: list, mcp_results: dict) -> dict:
    """
    整合 SKU 列表和 MCP 结果。
    - 只保留 sku_list 中存在的 SKU
    - 缺失的 SKU 添加为空数据结构
    
    输入格式: {sku: {"traffic_listing": {"items": [...], "data": {...}}}}
    输出格式: 保持相同
    """
    integrated = {}
    for sku in sku_list:
        if sku in mcp_results:
            integrated[sku] = mcp_results[sku]
        else:
            integrated[sku] = {"traffic_listing": {"items": [], "data": {}}}

    has_data = sum(1 for v in integrated.values() if v.get("traffic_listing", {}).get("items"))
    print(f"\n📊 整合完成: {has_data}/{len(integrated)} 个 SKU 有竞品数据")

    return integrated

def save_output(data: dict, output_dir: str) -> str:
    """保存到 _competitor_data.json"""
    os.makedirs(output_dir, exist_ok=True)
    output_path = os.path.join(output_dir, "_competitor_data.json")
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return output_path

# ── 主流程 ────────────────────────────────────────────────────────────────────

def main():
    # 获取默认 workspace（从环境变量或 ~/Desktop）
    default_workspace = os.getenv('WORKSPACE')
    if default_workspace:
        default_workspace = os.path.expanduser(default_workspace)
    else:
        default_workspace = os.path.expanduser('~/Desktop')

    parser = argparse.ArgumentParser(description="竞品数据整合")
    parser.add_argument("--mode", choices=["integrate", "collect"], default="integrate",
                        help="integrate: 整合已有MCP结果; collect: 直接调用MCP获取")
    parser.add_argument("--workspace", help="workspace 目录（新方式，会自动推导 source 和 output）")
    parser.add_argument("--date", help="数据日期 YYYYMMDD（与 --workspace 一起使用）")
    parser.add_argument("--source", help="复筛表路径（旧方式）")
    parser.add_argument("--mcp-results", help="MCP结果JSON路径（integrate模式必需）")
    parser.add_argument("--output", help="输出目录（旧方式）")
    parser.add_argument("--workers", type=int, default=1, help="并发数（已废弃，强制串行）")
    parser.add_argument("--limit", type=int, default=None,
                        help="限制处理的SKU数量（取前N个），用于测试")
    args = parser.parse_args()

    # ── 推断日期 ──
    date_str = datetime.now().strftime("%Y%m%d")

    # 处理新方式（--workspace + --date）
    if args.workspace and args.date:
        date_str = args.date
        # 自动推导 source 和 output
        workspace_path = os.path.expanduser(args.workspace)
        source_dir = os.path.join(workspace_path, f"{args.date}选品")
        args.source = os.path.join(source_dir, f"{args.date}选品复筛.xlsx")
        args.output = source_dir
        print(f"📁 使用 workspace 模式: {workspace_path}")
        print(f"   source: {args.source}")
        print(f"   output: {args.output}")
    # 处理仅 --date 的情况（使用默认 workspace）
    elif args.date and not args.workspace:
        date_str = args.date
        workspace_path = default_workspace
        source_dir = os.path.join(workspace_path, f"{args.date}选品")
        args.source = os.path.join(source_dir, f"{args.date}选品复筛.xlsx")
        args.output = source_dir
        print(f"📁 使用默认 workspace: {workspace_path}")
        print(f"   source: {args.source}")
        print(f"   output: {args.output}")

    # 旧方式日期推断
    if args.source:
        m = re.search(r'(\d{8})', args.source)
        if m:
            date_str = m.group(1)
    print(f"📅 日期: {date_str}")

    # ── 模式A: 整合模式（推荐）──
    if args.mode == "integrate":
        if not args.source or not args.mcp_results or not args.output:
            print("❌ integrate 模式需要: --source, --mcp-results, --output")
            sys.exit(1)

        if not os.path.exists(args.source):
            print(f"❌ 找不到复筛表: {args.source}")
            sys.exit(1)
        if not os.path.exists(args.mcp_results):
            print(f"❌ 找不到MCP结果文件: {args.mcp_results}")
            sys.exit(1)

        # 读取 SKU 列表
        sku_list, _, _ = read_skus_from_xlsx(args.source)
        
        # 限制处理的SKU数量
        if args.limit:
            sku_list = sku_list[:args.limit]
            print(f"🧪 测试模式: 只处理前 {len(sku_list)} 个 SKU")
        print(f"✅ 从复筛表读取 {len(sku_list)} 个 SKU")

        # 读取 MCP 结果
        with open(args.mcp_results, encoding='utf-8') as f:
            mcp_results = json.load(f)
        
        # 如果有limit，也限制mcp_results
        if args.limit:
            mcp_results = {k: v for i, (k, v) in enumerate(mcp_results.items()) if i < args.limit}
        
        print(f"✅ 读取 MCP 结果 {len(mcp_results)} 条")

        # 整合
        integrated = integrate_data(sku_list, mcp_results)

        # 输出
        output_path = save_output(integrated, args.output)
        print(f"\n✅ 竞品数据已保存: {output_path}")
        return 0

    # ── 模式B: 直接收集 ──
    if args.mode == "collect":
        if not args.source or not args.output:
            print("❌ collect 模式需要: --source, --output")
            sys.exit(1)

        if not os.path.exists(args.source):
            print(f"❌ 找不到复筛表: {args.source}")
            sys.exit(1)

        # 读取 SKU 列表及相关数据
        sku_list, price_map, node_id_map = read_skus_from_xlsx(args.source)
        print(f"✅ 从复筛表读取 {len(sku_list)} 个 SKU")
        print(f"   价格数据: {len(price_map)} 个, 类目ID: {len(node_id_map)} 个")

        # 检查 mcp CLI
        try:
            subprocess.run(["mcp", "--version"], capture_output=True, check=True)
        except (subprocess.CalledProcessError, FileNotFoundError):
            print("❌ 未找到 mcp CLI 或版本检查失败")
            print("   请确保 mcp 已安装: npm install -g @modelcontextprotocol/cli")
            sys.exit(1)

        competitor_data = prefetch_all(sku_list, price_map, node_id_map, max_workers=args.workers)
        output_path = save_output(competitor_data, args.output)
        print(f"\n✅ 竞品数据已保存: {output_path}")
        return 0

if __name__ == '__main__':
    sys.exit(main())
