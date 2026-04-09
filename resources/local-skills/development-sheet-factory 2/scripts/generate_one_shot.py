#!/usr/bin/env python3
"""
一键生成开发表格 —— development-sheet-factory 单步流程
=======================================================
支持两种运行模式：

## 批量模式（读取 mcp_results.json）
  python3 scripts/generate_one_shot.py --workspace ~/Desktop --date 20260407
  python3 scripts/generate_one_shot.py --date 20260407 --limit 3
  python3 scripts/generate_one_shot.py --date 20260407 --spu B0DLD8LM5T,B0DQJ6F4G2

## 单步模式（AI 逐个调用，每获取一个竞品就生成一个表格）
  python3 scripts/generate_one_shot.py --workspace ~/Desktop --date 20260407 \
      --single B0DLD8LM5T --competitor-json '{"traffic_listing":{"items":[{"asin":"B0xxx","price":9.99,"rating":4.5}]}}'

  --single <SPU_ID>              指定单个 SPU（必需）
  --competitor-json '<JSON>'     该 SPU 的竞品数据 JSON（可选，无则仅填充基础字段）

执行流程（单步模式）:
  1. 从复筛表读取该 SPU 的行数据
  2. 解析传入的竞品数据
  3. 下载主图
  4. 生成开发表格
  5. 将竞品数据追加写入 mcp_results.json
  6. 输出文件名，供 AI 确认

执行流程（批量模式）:
  1. 读取复筛表，提取 SPU 列表
  2. 读取 mcp_results.json 中的竞品数据
  3. 下载全部主图
  4. 逐个生成开发表格
  5. 输出 _READY.json
"""

import sys
import os
import json
import re
import glob
import shutil
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

try:
    import openpyxl
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
    from openpyxl.drawing.xdr import XDRPositiveSize2D
    from openpyxl.utils.units import pixels_to_EMU
    import urllib.request
    from PIL import Image as PILImage
except ImportError as e:
    print(f"缺少依赖: {e}\n请运行: pip3 install openpyxl Pillow")
    sys.exit(1)

# ── 常量 ──────────────────────────────────────────────────────────────────────

SKILL_DIR     = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE_PATH = os.path.join(SKILL_DIR, "assets", "开发模版.xlsx")
_PRICE_RE     = re.compile(r'[\d.]+')


# ── 工具函数 ──────────────────────────────────────────────────────────────────

def safe_filename(name: str) -> str:
    name = re.sub(r'[\\/:*?"<>|]', '_', name)
    return name[:50]

def parse_price(price_str):
    if not price_str:
        return None
    m = _PRICE_RE.search(str(price_str))
    return float(m.group()) if m else None

def download_image(url: str, tmp_path: str) -> bool:
    try:
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=20) as resp:
            with open(tmp_path, 'wb') as f:
                f.write(resp.read())
        return os.path.getsize(tmp_path) > 100
    except Exception:
        return False

def _calc_f2_size(ws) -> tuple:
    col_w_f = ws.column_dimensions['F'].width or 13
    col_w_g = ws.column_dimensions['G'].width or 13
    col_w_h = ws.column_dimensions['H'].width or 13
    total_col_w = col_w_f + col_w_g + col_w_h
    row_h = ws.row_dimensions[2].height or 147
    PX_PER_CHAR = 7
    PX_PER_PT   = 96 / 72
    return total_col_w * PX_PER_CHAR, row_h * PX_PER_PT

def _fit_image(img_path: str, max_w: float, max_h: float) -> tuple:
    with PILImage.open(img_path) as im:
        orig_w, orig_h = im.size
    ratio = min(max_w / orig_w, max_h / orig_h)
    return orig_w * ratio, orig_h * ratio


# ── 竞品数据解析 ──────────────────────────────────────────────────────────────

def parse_competitor(data: dict) -> dict:
    if not data:
        return {}
    if isinstance(data, list):
        data = {'traffic_listing': {'items': data, 'data': {}}}

    items = data.get('traffic_listing', {}).get('items', [])
    if not items:
        return {}

    valid = [it for it in items if it.get('asin')][:3]
    if not valid:
        return {}

    result = {}
    base_url = "https://www.amazon.de/dp/"

    if len(valid) >= 1:
        result['link1'] = base_url + valid[0]['asin']
    if len(valid) >= 2:
        result['link2'] = base_url + valid[1]['asin']

    all_items = [it for it in items if it.get('price')]
    if all_items:
        prices = [it.get('price') for it in all_items]
        lo, hi = min(prices), max(prices)
        result['price_range'] = f"€{lo:.2f}" if abs(hi - lo) < 0.01 else f"€{lo:.2f} - €{hi:.2f}"

    all_ratings = [it.get('rating') for it in items if it.get('rating')]
    if all_ratings:
        avg = round(sum(all_ratings) / len(all_ratings), 1)
        result['avg_rating'] = f"{avg}/5.0"
        low_count = sum(1 for r in all_ratings if r < 4.0)
        n = len(all_ratings)
        if low_count == 0:
            result['rating_summary'] = f"基于{n}个竞品，均分{avg}，整体口碑良好"
        else:
            result['rating_summary'] = f"基于{n}个竞品，均分{avg}，{low_count}个评分偏低"

    return result


# ── 单文件生成 ────────────────────────────────────────────────────────────────

def generate_sheet(row_data: dict, output_dir: str, img_path=None, competitor=None) -> str:
    sku_id        = row_data.get('SKU_ID') or row_data.get('SPU_ID', '')
    spu_id        = row_data.get('SPU_ID') or sku_id
    cn_title      = row_data.get('中文标题', '')
    link          = row_data.get('商品链接', '')
    image_url     = row_data.get('主图链接', '')
    price_str     = row_data.get('售价', '')
    is_seasonal   = row_data.get('季节性产品', '否')
    seasonal_month= row_data.get('节日月份', '')

    price    = parse_price(price_str)
    fname    = safe_filename(f"{spu_id}+{cn_title}") + ".xlsx"
    out_path = os.path.join(output_dir, fname)

    wb = openpyxl.load_workbook(TEMPLATE_PATH)
    ws = wb.active

    # 基础字段
    ws['C2'] = link
    ws['O2'] = is_seasonal
    ws['P2'] = seasonal_month or ''
    ws['B7'] = cn_title
    if price is not None:
        ws['I7'] = price
        ws['J7'] = price

    # 主图
    if img_path:
        try:
            cell_w, cell_h = _calc_f2_size(ws)
            target_w, target_h = _fit_image(img_path, cell_w, cell_h)
            img = XLImage(img_path)
            img.width  = target_w
            img.height = target_h
            marker = AnchorMarker(col=5, row=1, colOff=0, rowOff=0)
            size = XDRPositiveSize2D(cx=pixels_to_EMU(target_w), cy=pixels_to_EMU(target_h))
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            ws.add_image(img)
        except Exception:
            if image_url:
                ws['F2'] = image_url
    elif image_url:
        ws['F2'] = image_url

    # 竞品字段
    if competitor:
        cell_map = {
            'link1': 'D2', 'link2': 'E2', 'price_range': 'I2',
            'avg_rating': 'S2', 'rating_summary': 'T2',
        }
        for key, cell in cell_map.items():
            val = competitor.get(key, '')
            if val:
                ws[cell] = val

    wb.save(out_path)
    return out_path


# ── 增量写入 mcp_results ────────────────────────────────────────────────────

def load_mcp_results(mcp_path: str) -> dict:
    """加载已有的 mcp_results.json（可能不存在或为空）"""
    if os.path.exists(mcp_path):
        try:
            with open(mcp_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except (json.JSONDecodeError, IOError):
            return {}
    return {}

def save_mcp_results(mcp_path: str, data: dict):
    """保存 mcp_results.json"""
    with open(mcp_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


# ── 主流程 ────────────────────────────────────────────────────────────────────

def _read_source_rows(source_xlsx: str):
    """读取复筛表并 SPU 去重，返回行列表"""
    wb_src = openpyxl.load_workbook(source_xlsx)
    ws_src = wb_src.active
    headers = [cell.value for cell in ws_src[1]]
    rows = []
    for row in ws_src.iter_rows(min_row=2, values_only=True):
        if any(row):
            rows.append(dict(zip(headers, row)))
    seen_spu = set()
    spu_rows = []
    for row in rows:
        spu = row.get('SPU_ID') or row.get('SKU_ID', '')
        if spu not in seen_spu:
            seen_spu.add(spu)
            spu_rows.append(row)
    return spu_rows


def run_single(workspace: str, date_str: str, spu_id: str, competitor_json: str = None):
    """单步模式：处理单个 SPU，生成一个开发表格"""
    source_dir = os.path.join(workspace, f"{date_str}选品")
    source_xlsx = os.path.join(source_dir, f"{date_str}选品复筛.xlsx")
    output_dir = os.path.join(source_dir, f"{date_str}开发表格")
    mcp_results_path = os.path.join(source_dir, "mcp_results.json")

    if not os.path.exists(source_xlsx):
        print(f"ERROR: 找不到复筛表: {source_xlsx}")
        sys.exit(1)

    os.makedirs(output_dir, exist_ok=True)

    # 读取复筛表，找到该 SPU 的行
    all_rows = _read_source_rows(source_xlsx)
    row = None
    for r in all_rows:
        if (r.get('SPU_ID') or r.get('SKU_ID', '')) == spu_id:
            row = r
            break

    if not row:
        print(f"ERROR: 在复筛表中找不到 SPU: {spu_id}")
        sys.exit(1)

    title = row.get('中文标题', '')
    image_url = row.get('主图链接', '')

    # 解析竞品数据
    competitor = None
    comp_data = None
    if competitor_json:
        try:
            comp_data = json.loads(competitor_json)
            competitor = parse_competitor(comp_data)
        except json.JSONDecodeError as e:
            print(f"WARN: 竞品数据 JSON 解析失败: {e}")

    # 下载主图
    img_path = None
    if image_url:
        tmp_path = os.path.join("/tmp", f"_dev_img_{spu_id}.jpg")
        if download_image(image_url, tmp_path):
            img_path = tmp_path

    # 生成表格
    try:
        out = generate_sheet(row, output_dir, img_path=img_path, competitor=competitor)
        fname = os.path.basename(out)
        tag = " +竞品" if competitor else ""
        img_tag = " +主图" if img_path else ""
        print(f"OK: {spu_id}  {title[:25]}  -> {fname}{tag}{img_tag}")
    except Exception as e:
        print(f"ERROR: {spu_id}  -> {e}")
        sys.exit(1)

    # 追加写入 mcp_results.json
    if comp_data:
        mcp_results = load_mcp_results(mcp_results_path)
        mcp_results[spu_id] = comp_data
        save_mcp_results(mcp_results_path, mcp_results)

    # 输出结果供 AI 解析
    print(f"OUTPUT:{out}")


def run_batch(workspace: str, date_str: str, limit_count=None, filter_spus=None, no_competitor=False):
    """批量模式：读取 mcp_results.json，一次性生成全部开发表格"""
    source_dir = os.path.join(workspace, f"{date_str}选品")
    source_xlsx = os.path.join(source_dir, f"{date_str}选品复筛.xlsx")
    output_dir = os.path.join(source_dir, f"{date_str}开发表格")
    mcp_results_path = os.path.join(source_dir, "mcp_results.json")

    if not os.path.exists(source_xlsx):
        print(f"找不到复筛表: {source_xlsx}")
        sys.exit(1)
    if not os.path.exists(TEMPLATE_PATH):
        print(f"找不到模板: {TEMPLATE_PATH}")
        sys.exit(1)

    os.makedirs(output_dir, exist_ok=True)

    rows = _read_source_rows(source_xlsx)
    print(f"复筛表: {len(rows)} 个 SPU")

    if limit_count:
        rows = rows[:limit_count]
        print(f"测试模式: 只处理前 {len(rows)} 个")
    if filter_spus:
        original = len(rows)
        rows = [r for r in rows if (r.get('SPU_ID') or r.get('SKU_ID', '')) in filter_spus]
        print(f"增量模式: {len(rows)} 个指定 SPU（原 {original} 个）")

    mcp_results = {} if no_competitor else load_mcp_results(mcp_results_path)
    if mcp_results:
        print(f"已有竞品数据: {sum(1 for v in mcp_results.values() if v.get('traffic_listing', {}).get('items'))} 个 SPU")
    else:
        print("无竞品数据文件，将仅生成基础字段")

    # 并发预下载图片
    print(f"\n下载主图...")
    img_map = {}
    tasks = [(i, row.get('主图链接', '')) for i, row in enumerate(rows, 1)]
    with ThreadPoolExecutor(max_workers=8) as pool:
        futures = {pool.submit(_fetch_img, t): t[0] for t in tasks}
        for fut in as_completed(futures):
            idx, path = fut.result()
            img_map[idx] = path
    ok_count = sum(1 for v in img_map.values() if v)
    print(f"图片下载: {ok_count}/{len(tasks)} 成功")

    # 逐个生成开发表格
    print(f"\n生成开发表格 -> {output_dir}\n")
    generated = []
    failed = []
    comp_filled = 0

    for i, row in enumerate(rows, 1):
        spu = row.get('SPU_ID') or row.get('SKU_ID', '?')
        sku = row.get('SKU_ID') or spu
        title = row.get('中文标题', '')

        comp_data = mcp_results.get(sku) or mcp_results.get(spu)
        competitor = None
        if comp_data:
            competitor = parse_competitor(comp_data)

        try:
            path = generate_sheet(row, output_dir, img_path=img_map.get(i), competitor=competitor)
            generated.append({"sku": spu, "file": os.path.basename(path)})
            tag = " +竞品" if competitor else ""
            if competitor:
                comp_filled += 1
            print(f"  [{i:02d}/{len(rows)}] {spu}  {title[:20]}  -> {os.path.basename(path)}{tag}")
        except Exception as e:
            print(f"  [{i:02d}/{len(rows)}] {spu}  -> 失败: {e}")
            failed.append({"sku": spu, "error": str(e)})

    # 写 _READY.json
    ready = {
        "schema_version": "1.0",
        "date": date_str,
        "workspace": workspace,
        "status": "ready" if not failed else "partial",
        "total_spu": len(rows),
        "generated": len(generated),
        "failed": len(failed),
        "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "template_path": TEMPLATE_PATH,
        "competitor_filled": comp_filled,
        "competitor_skipped": len(generated) - comp_filled,
        "files": generated,
        "errors": failed,
    }
    ready_path = os.path.join(output_dir, "_READY.json")
    with open(ready_path, 'w', encoding='utf-8') as f:
        json.dump(ready, f, ensure_ascii=False, indent=2)

    print(f"\n{'='*50}")
    print(f"完成: {len(generated)} 个文件生成成功，{len(failed)} 个失败")
    print(f"竞品数据: {comp_filled} 个已填写，{len(generated) - comp_filled} 个无数据")
    print(f"输出目录: {output_dir}")
    if failed:
        print(f"失败列表: {[x['sku'] for x in failed]}")


def _fetch_img(task):
    idx, url = task
    if not url:
        return idx, None
    tmp_path = os.path.join("/tmp", f"_dev_img_{idx}.jpg")
    ok = download_image(url, tmp_path)
    return idx, tmp_path if ok else None


def main():
    args = sys.argv[1:]
    if not args:
        print("用法:")
        print("  # 批量模式")
        print("  python3 generate_one_shot.py --date <YYYYMMDD>")
        print("  python3 generate_one_shot.py --workspace <DIR> --date <YYYYMMDD>")
        print("  python3 generate_one_shot.py --date <YYYYMMDD> --limit 3")
        print("  python3 generate_one_shot.py --date <YYYYMMDD> --spu B0DLD8LM5T,B0DQJ6F4G2")
        print("")
        print("  # 单步模式（AI 逐个调用）")
        print("  python3 generate_one_shot.py --date <YYYYMMDD> --single <SPU_ID> --competitor-json '<JSON>'")
        print("")
        print("参数:")
        print("  --workspace <DIR>              workspace 目录（默认 ~/Desktop）")
        print("  --date <YYYYMMDD>              数据日期（必需）")
        print("  --single <SPU_ID>              单步模式，处理单个 SPU")
        print("  --competitor-json '<JSON>'     单步模式的竞品数据 JSON")
        print("  --limit <N>                    批量模式，只处理前 N 个")
        print("  --spu <ID1,ID2>                批量模式，指定 SPU 列表")
        print("  --no-competitor                跳过竞品数据，仅生成基础字段")
        sys.exit(1)

    # 参数解析
    default_workspace = os.environ.get('WORKSPACE', os.path.expanduser('~/Desktop'))
    workspace = default_workspace
    date_str = None
    limit_count = None
    filter_spus = None
    no_competitor = False
    single_spu = None
    competitor_json = None

    i = 0
    while i < len(args):
        if args[i] == '--workspace' and i + 1 < len(args):
            workspace = args[i + 1]
            i += 2
        elif args[i] == '--date' and i + 1 < len(args):
            date_str = args[i + 1]
            i += 2
        elif args[i] == '--limit' and i + 1 < len(args):
            limit_count = int(args[i + 1])
            i += 2
        elif args[i] == '--spu' and i + 1 < len(args):
            filter_spus = set(s.strip() for s in args[i + 1].split(',') if s.strip())
            i += 2
        elif args[i] == '--no-competitor':
            no_competitor = True
            i += 1
        elif args[i] == '--single' and i + 1 < len(args):
            single_spu = args[i + 1].strip()
            i += 2
        elif args[i] == '--competitor-json' and i + 1 < len(args):
            competitor_json = args[i + 1]
            i += 2
        else:
            i += 1

    if not date_str:
        print("缺少必要参数: --date")
        sys.exit(1)

    workspace = os.path.expanduser(workspace)

    # 单步模式 vs 批量模式
    if single_spu:
        run_single(workspace, date_str, single_spu, competitor_json)
    else:
        run_batch(workspace, date_str, limit_count, filter_spus, no_competitor)


if __name__ == '__main__':
    main()
